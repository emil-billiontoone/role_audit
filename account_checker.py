"""
Excel Parser Module

A utility for parsing Excel (.xlsx) files into lists of key-value pairs.
Handles employee data with multiple columns and data types.
"""

import os
import sys
import argparse
from typing import List, Dict, Any
import s4
import s4.clarity
import keyring
import traceback
from openpyxl import load_workbook


class ExcelParser:
    """
    A class for parsing Excel (.xlsx) files into key-value pair dictionaries.
    """

    def __init__(self, sheet_name: str = None):
        """
        Initialize the Excel parser.

        Args:
            sheet_name: Name of the sheet to parse (default: active sheet)
        """
        self.sheet_name = sheet_name

    def parse_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse an Excel (.xlsx) file and return a list of dictionaries.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of dictionaries with column headers as keys

        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file is empty or malformed
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            # Load the workbook
            workbook = load_workbook(filename=file_path, data_only=True)

            # Select the worksheet
            if self.sheet_name:
                if self.sheet_name in workbook.sheetnames:
                    worksheet = workbook[self.sheet_name]
                else:
                    raise ValueError(f"Sheet '{self.sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            else:
                worksheet = workbook.active

            # Get all rows as a list
            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                raise ValueError("Excel worksheet is empty")

            # Use first row as headers
            headers = [str(cell) if cell is not None else f"column_{i+1}" for i, cell in enumerate(rows[0])]
            data_rows = rows[1:]

            # Convert each row to a dictionary
            result = []
            for row_num, row in enumerate(data_rows, start=2):  # Start at 2 for line numbering
                try:
                    # Handle rows with different lengths
                    row_dict = {}
                    for i, header in enumerate(headers):
                        value = row[i] if i < len(row) else None
                        # Clean and process the value
                        row_dict[header.strip()] = self._process_value(value)

                    # Skip completely empty rows
                    if any(value for value in row_dict.values()):
                        result.append(row_dict)
                except Exception as e:
                    print(f"Warning: Error processing row {row_num}: {e}")
                    continue

            workbook.close()
            return result

        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {e}")

    def _process_value(self, value: Any) -> Any:
        """
        Process and clean individual cell values from Excel.

        Args:
            value: Raw cell value (can be string, number, date, etc.)

        Returns:
            Processed value (string, int, float, or cleaned string)
        """
        if value is None:
            return ""

        # Handle different Excel data types
        if isinstance(value, (int, float)):
            return value

        if isinstance(value, str):
            value = value.strip()
            if not value:
                return ""
            return value

        # Handle dates and other types by converting to string
        return str(value)


def main():
    """
    Command line interface for the Excel parser.
    """
    parser = argparse.ArgumentParser(
        description="Parse Excel (.xlsx) files into key-value pairs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python account_checker.py data.xlsx
        """,
    )

    parser.add_argument("filename", help="Excel (.xlsx) file to parse")
    parser.add_argument("--unknown", "-u", action="store_true", help="Manually process unknown accounts")

    args = parser.parse_args()

    # Check if file exists
    if not os.path.exists(args.filename):
        print(f"Error: File '{args.filename}' not found.")
        sys.exit(1)

    # Create Excel parser
    excel_parser = ExcelParser()

    try:

        # Parse the Excel file
        print(f"Parsing '{args.filename}'...")

        data = excel_parser.parse_file(args.filename)

        print(f"Successfully loaded {len(data)} records")

        account_data = {}
        added_to_account_data = 0
        duplicate_accounts = 0

        for record in data:
            email = record.get("Work Contact: Work Email")
            name = f"{record.get('Legal First Name', '')} {record.get('Legal Last Name', '')}".strip()

            current_record = None

            if name in account_data:
                current_record = account_data[name]
            elif email and email in account_data:
                current_record = account_data[email]

            if current_record:
                # print(f"Duplicate name found: {name}")
                if current_record["Position Status"] == "Active":
                    # print(f"Don't update data {current_record['Position Status']} > {record['Position Status']}")
                    duplicate_accounts += 1
                    continue
                # print(f"Updating record {current_record['Position Status']} with {record['Position Status']}")

            account_data[name] = record
            if email:
                account_data[email] = record
            added_to_account_data += 1

        # Save account data to text file in same directory as the input file
        # output_dir = os.path.dirname(args.filename)
        # output_filename = os.path.splitext(os.path.basename(args.filename))[0] + "_account_data.txt"
        # output_path = os.path.join(output_dir, output_filename)

        # with open(output_path, "w", encoding="utf-8") as f:
        #     f.write(json.dumps(account_data))

        # print(f"Account data saved to: {output_path}")

        print(f"Records Found: {len(data)}")
        print(f"Records Indexed by Name/Work Email: {added_to_account_data} ({duplicate_accounts} duplicates skipped)")

        # dev, staging, or prod
        server = "dev"

        # account settings
        username = "bgriffiths"
        password = keyring.get_password(f"clarity-{server}", username)

        # lims object
        lims = s4.clarity.LIMS(f"https://clarity-{server}.btolims.com/api/v2/", username, password)
        print(f'API version: {lims.versions[0]["major"]}')

        # load all accounts
        print("Loading Clarity accounts... (this can take a while)")
        accounts = lims.researchers.all()

        print(f"Found {len(accounts)} Clarity accounts")
        limited_role = lims.roles.get_by_name("Limited (BTO)")
        print(limited_role.limsid, limited_role.name)

        clarity_accounts_unknown = []
        clarity_accounts_stripped = []
        clarity_accounts_deactivated = []
        clarity_accounts_active = []
        clarity_accounts_archived = []
        clarity_accounts_ignored = []

        print("Processing Clarity accounts... (this can take a while)")
        for account in accounts:
            account.invalidate()
            if account.get("BTO Legal Name", "") in ["ignore", "Terminated"]:
                clarity_accounts_ignored.append(account)
                continue

            employee = None
            name = f"{account.first_name} {account.last_name}"
            if name in account_data:
                employee = account_data[name]
            elif account.email in account_data:
                employee = account_data[account.email]
            elif account.get("BTO Legal Name", None) in account_data:
                employee = account_data[account.get("BTO Legal Name")]

            if employee == None:
                clarity_accounts_unknown.append(account)
                continue

            employee_status = employee.get("Position Status", "")

            if employee_status in ["Active", "Leave"]:
                clarity_accounts_active.append(account)
                continue

            # print(name, employee_status)
            if employee_status == "Terminated" or account.locked or account.get("BTO Legal Name", "") == "Terminated":
                if len(account.roles) > 1 or not limited_role in account.roles:
                    for role in account.roles:
                        account.remove_role(role)
                    account.add_role(limited_role)
                    # print("\tStripped Roles")
                    clarity_accounts_stripped.append(account)
                if not account.locked:
                    account.locked = True
                    # print("\tArchived Account")
                    clarity_accounts_deactivated.append(account)

                if account.locked:
                    clarity_accounts_archived.append(account)
                account.commit()
                # it doesn't matter that we don't know if an archived account belongs to someebody.
                if account in clarity_accounts_unknown:
                    clarity_accounts_unknown.remove(account)
            else:
                raise Exception(f"Unknown Status: {employee['Position Status']}")

        print(f"Stripped roles from {len(clarity_accounts_stripped)} terminiated accounts")
        print(f"Archived {len(clarity_accounts_deactivated)} terminiated accounts")
        print(f"Remaining active accounts: {len(clarity_accounts_active)}")
        print(f"Total Archived Accounts: {len(clarity_accounts_archived)}")
        print(f"Total Ignored Accounts: {len(clarity_accounts_ignored)}")
        print(f"Unknown accounts: {len(clarity_accounts_unknown)}")

        for account in clarity_accounts_unknown:
            name = f"{account.first_name} {account.last_name}"
            if "BTO Legal Name" in account:
                print(f"Unknown account ({name}) with BTO Legal Name: {account['BTO Legal Name']}")
                continue

            print(f"Unknown account: {name} ({account.limsid})")
            print("\t0. Terminated\n\n\tPossible match(es) found:")

            if args.unknown:
                options = []
                options.append("Terminated")
                i = 0
                for record in data:
                    if (
                        record.get("Legal First Name", "") == account.first_name
                        or record.get("Legal Last Name", "") == account.last_name
                    ):
                        i += 1
                        print(
                            f"\t{i}. {record.get('Legal First Name', '')} {record.get('Legal Last Name', '')} - {record.get('Work Contact: Work Email', '')}"
                        )
                        options.append(f"{record.get('Legal First Name', '')} {record.get('Legal Last Name', '')}")
                choice = input(f"Select option 0–{len(options)-1}: ").strip()
                if choice.isdigit() and 0 <= int(choice) <= len(options) - 1:
                    print(f"{name} = {choice}. {options[int(choice)]}")
                    account["BTO Legal Name"] = options[int(choice)]
                    account.commit()
                    clarity_accounts_unknown.remove(account)
                    print("Clarity updated with BTO Legal Name")
                else:
                    print("Skipped")

            print(f"Finished processing Clarity accounts. {len(clarity_accounts_unknown)} unknown accounts remaining.")
        else:
            print("Finished processing Clarity accounts. All good.")

    except Exception as e:
        tb = traceback.format_exc()
        print(f"Error parsing Excel file: {e} {tb}")
        sys.exit(1)


if __name__ == "__main__":
    main()
